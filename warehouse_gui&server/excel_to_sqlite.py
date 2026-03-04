#!/usr/bin/env python3
"""
Excel 데이터를 SQLite DB로 변환
- 데이터_베이스.xlsx → warehouse.db
"""
import sqlite3
from openpyxl import load_workbook
import os

def create_database(db_path='warehouse.db'):
    """SQLite 데이터베이스 및 테이블 생성"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 기존 테이블 삭제 (초기화)
    cursor.execute("DROP TABLE IF EXISTS inventory")
    
    # 재고 테이블 생성
    cursor.execute("""
        CREATE TABLE inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shelf_number TEXT NOT NULL UNIQUE,
            item TEXT NOT NULL UNIQUE,
            stock INTEGER NOT NULL DEFAULT 0,
            zone INTEGER,
            row INTEGER,
            tier INTEGER,
            CHECK (stock >= 0)
        )
    """)
    
    # 인덱스 생성 (빠른 검색)
    cursor.execute("CREATE INDEX idx_item ON inventory(item)")
    cursor.execute("CREATE INDEX idx_shelf ON inventory(shelf_number)")
    
    conn.commit()
    print("✅ 데이터베이스 테이블 생성 완료")
    return conn

def parse_shelf_number(shelf_str):
    """선반번호 파싱: '1-2-3' → (1, 2, 3)"""
    try:
        parts = shelf_str.split('-')
        return int(parts[0]), int(parts[1]), int(parts[2])
    except:
        return None, None, None

def import_from_excel(conn, excel_path='데이터_베이스.xlsx'):
    """Excel 파일에서 데이터 읽어서 SQLite에 삽입"""
    cursor = conn.cursor()
    
    if not os.path.exists(excel_path):
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        return
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 스킵
        shelf_number = row[1]
        item = row[2]
        stock = row[3]
        
        if shelf_number and item and stock is not None:
            zone, row_num, tier = parse_shelf_number(shelf_number)
            
            cursor.execute("""
                INSERT INTO inventory (shelf_number, item, stock, zone, row, tier)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (shelf_number, item, stock, zone, row_num, tier))
            count += 1
    
    conn.commit()
    print(f"✅ {count}개 항목 임포트 완료")

def verify_data(conn):
    """데이터 검증"""
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM inventory")
    total = cursor.fetchone()[0]
    
    cursor.execute("SELECT zone, COUNT(*) FROM inventory GROUP BY zone ORDER BY zone")
    zones = cursor.fetchall()
    
    print(f"\n📊 데이터 검증")
    print(f"총 항목 수: {total}")
    print(f"구역별 분포:")
    for zone, count in zones:
        print(f"  구역 {zone}: {count}개")
    
    # 샘플 데이터 출력
    print(f"\n📦 샘플 데이터 (처음 5개):")
    cursor.execute("SELECT shelf_number, item, stock FROM inventory LIMIT 5")
    for shelf, item, stock in cursor.fetchall():
        print(f"  {shelf}: {item} ({stock}개)")

def main():
    print("="*60)
    print("Excel → SQLite 변환 시작")
    print("="*60)
    
    # 1. 데이터베이스 생성
    conn = create_database('warehouse.db')
    
    # 2. Excel 데이터 임포트
    import_from_excel(conn, '데이터_베이스.xlsx')
    
    # 3. 데이터 검증
    verify_data(conn)
    
    conn.close()
    print("\n✅ 변환 완료! warehouse.db 파일이 생성되었습니다.")

if __name__ == '__main__':
    main()