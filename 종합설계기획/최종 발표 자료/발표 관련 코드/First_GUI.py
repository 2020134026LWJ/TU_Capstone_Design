# 'warehouse_gui.py' 파일 전체 내용
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.core.text import LabelBase
from kivy.graphics import Color, Rectangle
from kivy.clock import Clock
from openpyxl import load_workbook
import os

# 윈도우 크기 설정 (라즈베리파이 5인치 터치스크린)
Window.size = (800, 480)

# 한글 폰트 등록
FONT_NAME = 'NanumGothic'

# Windows의 기본 한글 폰트 경로들
FONT_PATHS = [
    'C:/Windows/Fonts/malgun.ttf',  # 맑은 고딕
    'C:/Windows/Fonts/gulim.ttc',  # 굴림
    'C:/Windows/Fonts/batang.ttc',  # 바탕
    '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',  # Linux
    '/System/Library/Fonts/AppleGothic.ttf',  # Mac
]

# 사용 가능한 폰트 찾기
font_path = None
for path in FONT_PATHS:
    if os.path.exists(path):
        font_path = path
        break

if font_path:
    LabelBase.register(name=FONT_NAME, fn_regular=font_path)
else:
    print("경고: 한글 폰트를 찾을 수 없습니다. 기본 폰트를 사용합니다.")
    FONT_NAME = 'Roboto'  # Kivy 기본 폰트

# 엑셀 파일에서 물류 데이터 읽기
def load_logistics_data(excel_file='물류계획.xlsx'):
    """엑셀 파일에서 물류 데이터 로드"""
    logistics_data = {}
    work_list = {}
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 헤더 건너뛰고 데이터 읽기 (2행부터)
        for row in ws.iter_rows(min_row=2, values_only=True):
            manager = str(row[0])  # 담당자
            task_num = row[1]      # 작업번호
            locations = list(row[2:10])  # 위치1~8
            
            # 빈 셀 제거
            locations = [loc for loc in locations if loc]
            
            # logistics_data 구조 생성
            if manager not in logistics_data:
                logistics_data[manager] = []
            
            logistics_data[manager].append(locations)
        
        # work_list 생성 (작업물 버튼용)
        for manager, tasks in logistics_data.items():
            work_text = f"담당 {manager} 작업 목록:\n"
            for idx, task in enumerate(tasks, 1):
                work_text += f"{idx}번: " + ", ".join(task) + "\n"
            work_list[manager] = work_text
        
        print(f"✅ 엑셀 데이터 로드 완료: {excel_file}")
        print(f"📦 담당자 수: {len(logistics_data)}")
        for manager, tasks in logistics_data.items():
            print(f"   담당 {manager}: {len(tasks)}개 작업")
        
    except FileNotFoundError:
        print(f"⚠️ 엑셀 파일을 찾을 수 없습니다: {excel_file}")
        print("⚠️ 예시 데이터를 사용합니다.")
        # 예시 데이터
        logistics_data = {
            '1': [
                ['A(1-1)', 'B(2-2)', 'C(2-2)', 'D(3-1)', 'E(1-1)', 'F(1-4)', 'G(2-5)', 'H(4-1)'],
                ['I(2-3)', 'J(3-2)', 'K(1-5)', 'L(2-1)', 'M(4-2)', 'N(1-3)', 'O(3-4)', 'P(2-6)']
            ],
            '2': [
                ['Q(1-2)', 'R(2-4)', 'S(3-3)', 'T(1-6)', 'U(4-3)', 'V(2-1)', 'W(3-5)', 'X(1-4)']
            ]
        }
        work_list = {
            '1': "담당 1 작업 목록:\n1번: A(1-1), B(2-2)...\n2번: I(2-3)...",
            '2': "담당 2 작업 목록:\n1번: Q(1-2), R(2-4)..."
        }
    
    return logistics_data, work_list

# 전역 데이터 로드
LOGISTICS_DATA, WORK_LIST = load_logistics_data()


class WorkCell(Button):
    """작업 셀 (2x4 그리드의 각 칸)"""
    def __init__(self, **kwargs):
        super(WorkCell, self).__init__(**kwargs)
        # 배경색을 흰색이 아닌 연한 회색으로 변경하여 흰색 배경과 구분되게 함
        self.background_color = (0.9, 0.9, 0.9, 0.3)  
        self.color = (1, 1, 1, 1)  # 흰 텍스트
        self.font_size = '19sp'
        self.font_name = FONT_NAME
        self.is_completed = False
        
    def on_press(self):
        """셀 클릭시 빨간색으로 변경"""
        if not self.is_completed:
            self.background_color = (1, 0, 0, 0.5)  # 빨간색
            self.is_completed = True
        else:
            # 다시 클릭하면 연한 회색으로 돌아가도록 수정
            self.background_color = (0.9, 0.9, 0.9, 1)  
            self.is_completed = False


class WarehouseGUI(BoxLayout):
    def __init__(self, **kwargs):
        super(WarehouseGUI, self).__init__(**kwargs)
        
        # --- 전체 화면 배경색을 흰색으로 설정 ---
        with self.canvas.before:
            Color(0.9, 0.9, 0.9, 0.9)  # 흰색 (R, G, B, A)
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self.on_size_change, pos=self.on_size_change)
        # ------------------------------------
        
        self.orientation = 'vertical'
        self.padding = 10
        self.spacing = 10
        
        self.selected_manager = None
        self.current_task_index = 0
        self.work_cells = []
        self.task_number_labels = []
        
        # 상단 레이아웃
        self.build_top_section()
        
        # 중앙 작업 그리드 (2x4)
        self.build_work_grid()
        
        # 하단 완료 버튼
        self.build_bottom_section()
    
    def on_size_change(self, instance, value):
        """창 크기 변경 시 배경 사각형 업데이트"""
        if hasattr(self, 'rect'):
            self.rect.size = self.size
            self.rect.pos = self.pos

    def build_top_section(self):
        """상단 섹션: 담당자 선택, 선택된 담당자 표시, 작업물/작업시작 버튼"""
        top_layout = BoxLayout(orientation='horizontal', size_hint_y=0.25, spacing=10)
        
        # 왼쪽: 담당자 리스트 (스크롤 가능)
        manager_box = BoxLayout(orientation='vertical', size_hint_x=0.3)
        manager_label = Label(text='담당 선택', size_hint_y=0.2, font_size='16sp', font_name=FONT_NAME, color=(0,0,0,1)) # 글자색 검정
        manager_box.add_widget(manager_label)
        
        scroll_view = ScrollView(size_hint=(1, 0.8))
        manager_list = GridLayout(cols=1, spacing=5, size_hint_y=None)
        manager_list.bind(minimum_height=manager_list.setter('height'))
        
        # 담당자 버튼들 생성 (엑셀 데이터 기반)
        for manager_id in sorted(LOGISTICS_DATA.keys()):
            btn = Button(
                text=str(manager_id),
                size_hint_y=None,
                height=60,
                font_size='20sp',
                font_name=FONT_NAME,
                background_color=(0.3, 0.6, 1, 1), # 파란색 계열
                color=(1,1,1,1) # 글자색 흰색
            )
            btn.bind(on_press=self.on_manager_select)
            manager_list.add_widget(btn)
        
        scroll_view.add_widget(manager_list)
        manager_box.add_widget(scroll_view)
        top_layout.add_widget(manager_box)
        
        # 중앙: 선택된 담당자 표시
        self.selected_label = Label(
            text='',
            font_size='48sp',
            font_name=FONT_NAME,
            bold=True,
            size_hint_x=0.4,
            color=(0,0,0,1) # 글자색 검정
        )
        top_layout.add_widget(self.selected_label)
        
        # 오른쪽: 작업물, 작업시작 버튼
        right_box = BoxLayout(orientation='vertical', size_hint_x=0.3, spacing=10)
        
        self.work_list_btn = Button(
            text='작업물',
            font_size='20sp',
            font_name=FONT_NAME,
            background_color=(0.5, 0.5, 0.5, 0.5), # 연한 회색
            color=(1, 1, 1, 1), # 검은색 텍스트
            disabled=True
        )
        self.work_list_btn.bind(on_press=self.show_work_list)
        right_box.add_widget(self.work_list_btn)
        
        self.start_work_btn = Button(
            text='작업시작',
            font_size='20sp',
            font_name=FONT_NAME,
            background_color=(0.2, 0.8, 0.2, 1), # 녹색
            color=(1,1,1,1), # 글자색 흰색
            disabled=True
        )
        self.start_work_btn.bind(on_press=self.start_work)
        right_box.add_widget(self.start_work_btn)
        
        top_layout.add_widget(right_box)
        self.add_widget(top_layout)
        
    def build_work_grid(self):
        """중앙 작업 그리드 (2행 x 4열) with 작업 번호"""
        grid_container = BoxLayout(orientation='vertical', size_hint_y=0.6)
        grid_label = Label(
            text='작업 그리드',
            size_hint_y=0.1,
            font_size='18sp',
            font_name=FONT_NAME,
            bold=True,
            color=(0,0,0,1) # 글자색 검정
        )
        grid_container.add_widget(grid_label)
        
        # 메인 그리드 컨테이너 (번호 + 작업 셀)
        main_grid_layout = BoxLayout(orientation='horizontal', size_hint_y=0.9, spacing=10)
        
        # 왼쪽: 작업 번호 표시
        number_box = BoxLayout(orientation='vertical', size_hint_x=0.1, spacing=5)
        self.task_number_labels = []
        for i in range(2):  # 2행
            task_num_label = Label(
                text='',
                font_size='24sp',
                font_name=FONT_NAME,
                bold=True,
                color=(0, 0.5, 1, 1) # 파란색
            )
            self.task_number_labels.append(task_num_label)
            number_box.add_widget(task_num_label)
        
        main_grid_layout.add_widget(number_box)
        
        # 오른쪽: 작업 셀 그리드
        self.work_grid = GridLayout(cols=4, rows=2, spacing=5, size_hint_x=0.9)
        
        # 2x4 = 8개의 셀 생성
        for i in range(8):
            cell = WorkCell(text='')
            self.work_cells.append(cell)
            self.work_grid.add_widget(cell)
        
        main_grid_layout.add_widget(self.work_grid)
        grid_container.add_widget(main_grid_layout)
        self.add_widget(grid_container)
        
    def build_bottom_section(self):
        """하단 완료 버튼"""
        bottom_layout = BoxLayout(orientation='horizontal', size_hint_y=0.15, padding=10)
        
        self.complete_btn = Button(
            text='완료',
            font_size='24sp',
            font_name=FONT_NAME,
            background_color=(0, 0.5, 1, 1), # 파란색
            color=(1,1,1,1), # 글자색 흰색
            disabled=True
        )
        self.complete_btn.bind(on_press=self.complete_task)
        bottom_layout.add_widget(self.complete_btn)
        
        self.add_widget(bottom_layout)
        
    def on_manager_select(self, instance):
        """담당자 선택"""
        self.selected_manager = instance.text
        self.selected_label.text = instance.text
        self.current_task_index = 0
        
        # 버튼 활성화
        self.work_list_btn.disabled = False
        self.work_list_btn.background_color = (0.9, 0.9, 0.9, 1) # 활성화 시 연한 회색
        self.start_work_btn.disabled = False
        self.start_work_btn.background_color = (0.2, 0.8, 0.2, 1) # 활성화 시 녹색
        
        # 그리드 초기화
        for cell in self.work_cells:
            cell.text = ''
            cell.background_color = (0.9, 0.9, 0.9, 1) # 연한 회색으로 초기화
            cell.is_completed = False
        
        # 작업 번호 초기화
        for label in self.task_number_labels:
            label.text = ''
            
        self.complete_btn.disabled = True
        self.complete_btn.background_color = (0.5, 0.5, 0.5, 1) # 비활성화 시 어두운 회색 (변경)
        
    def show_work_list(self, instance):
        """작업물 버튼 클릭 - 작업 목록 팝업"""
        if self.selected_manager and self.selected_manager in WORK_LIST:
            content = BoxLayout(orientation='vertical', padding=10, spacing=10)

            with content.canvas.before:
                Color(1, 1, 1, 1)  # 흰색 (R, G, B, A)
                # 'content' 위젯 자체에 사각형을 바인딩합니다.
                content.rect = Rectangle(size=content.size, pos=content.pos)

            # 팝업 크기 변경 시 배경도 같이 변경되도록 바인딩
            def update_rect(instance, value):
                if hasattr(instance, 'rect'):
                    instance.rect.pos = instance.pos
                    instance.rect.size = instance.size
            content.bind(pos=update_rect, size=update_rect)
            # ------------------------------------
            
            work_text = Label(
                text=WORK_LIST[self.selected_manager],
                font_size='16sp',
                font_name=FONT_NAME,
                halign='left',
                valign='top',
                color=(0,0,0,1) # 글자색 검정 (흰 배경과 대비되어 잘 보입니다)
            )
            work_text.bind(size=work_text.setter('text_size'))
            
            scroll = ScrollView()
            scroll.add_widget(work_text)
            content.add_widget(scroll)
            
            close_btn = Button(
                text='닫기',
                size_hint_y=0.2,
                font_size='18sp',
                font_name=FONT_NAME,
                background_color=(0.7,0.7,0.7,1), # 닫기 버튼 색상 변경
                color=(0,0,0,1) # 글자색 검정
            )
            content.add_widget(close_btn)
            
            popup = Popup(
                title='작업 목록',
                content=content,
                size_hint=(0.8, 0.8),
                title_color=(0,0,0,1), # 팝업 타이틀 글자색 검정
            )
            close_btn.bind(on_press=popup.dismiss)
            popup.open()
            
    def start_work(self, instance):
        """작업 시작 - 첫 번째 물류 계획을 그리드에 표시"""
        if self.selected_manager and self.selected_manager in LOGISTICS_DATA:
            self.load_task_to_grid()
            self.complete_btn.disabled = False
            self.complete_btn.background_color = (0, 0.7, 1, 1) # 활성화 시 밝은 파란색
            
    def load_task_to_grid(self):
        """현재 작업을 그리드에 로드"""
        if self.selected_manager and self.selected_manager in LOGISTICS_DATA:
            tasks = LOGISTICS_DATA[self.selected_manager]
            
            if self.current_task_index < len(tasks):
                current_task = tasks[self.current_task_index]
                
                # 작업 번호 업데이트
                task_display_number = self.current_task_index + 1
                for i, label in enumerate(self.task_number_labels):
                    if i == 0:
                        label.text = f'{task_display_number}번'
                    else:
                        label.text = ''
                
                # 그리드 초기화 및 새 작업 로드
                for i, cell in enumerate(self.work_cells):
                    if i < len(current_task):
                        cell.text = current_task[i]
                        cell.background_color = (0.93, 0.93, 0.93, 1) # 연한 회색으로 초기화
                        cell.is_completed = False
                    else:
                        cell.text = ''
                        cell.background_color = (0.97, 0.97, 0.97, 1) # 빈 셀은 더 연한 회색으로
            else:
                # 모든 작업 완료
                self.show_completion_message()
                
    def complete_task(self, instance):
        """완료 버튼 - 모든 셀이 완료되었는지 확인 후 다음 물류 계획으로 이동"""
        # 내용이 있는 셀 중 완료되지 않은 셀 찾기
        incomplete_cells = []
        for cell in self.work_cells:
            if cell.text and not cell.is_completed:  # 텍스트가 있는데 완료 안됨
                incomplete_cells.append(cell)
        
        if incomplete_cells:
            # 미완료 셀이 있으면 노란색으로 5번 깜빡임
            self.blink_cells(incomplete_cells, 5)
        else:
            # 모든 셀이 완료되었으면 다음 작업으로
            self.current_task_index += 1
            self.load_task_to_grid()
    
    def blink_cells(self, cells, count):
        """셀들을 노란색으로 깜빡이게 하기"""
        blink_state = {'count': 0, 'max_count': count * 2, 'is_yellow': False}
        
        def toggle_color(dt):
            if blink_state['count'] >= blink_state['max_count']:
                # 깜빡임 완료 - 원래 색상(연한 회색)으로 복원
                for cell in cells:
                    cell.background_color = (0.9, 0.9, 0.9, 1) 
                return False  # 스케줄 중단
            
            # 노란색 <-> 연한 회색 토글
            if blink_state['is_yellow']:
                for cell in cells:
                    cell.background_color = (0.9, 0.9, 0.9, 1)  # 연한 회색
                blink_state['is_yellow'] = False
            else:
                for cell in cells:
                    cell.background_color = (1, 1, 0, 1)  # 노란색
                blink_state['is_yellow'] = True
            
            blink_state['count'] += 1
            return True  # 계속 반복
        
        # 0.3초마다 색상 토글
        Clock.schedule_interval(toggle_color, 0.3)
        
    def show_completion_message(self):
        """모든 작업 완료 메시지"""
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        msg = Label(
            text=f'담당 {self.selected_manager}의\n모든 작업이 완료되었습니다!',
            font_size='20sp',
            font_name=FONT_NAME,
            halign='center',
            color=(0,0,0,1) # 글자색 검정
        )
        content.add_widget(msg)
        
        close_btn = Button(
            text='확인',
            size_hint_y=0.3,
            font_size='18sp',
            font_name=FONT_NAME,
            background_color=(0.7,0.7,0.7,1), # 확인 버튼 색상 변경
            color=(0,0,0,1) # 글자색 검정
        )
        content.add_widget(close_btn)
        
        popup = Popup(
            title='작업 완료',
            content=content,
            size_hint=(0.6, 0.4),
            title_color=(0,0,0,1), # 팝업 타이틀 글자색 검정
        )
        close_btn.bind(on_press=popup.dismiss)
        popup.open()
        
        # 그리드 비우기
        for cell in self.work_cells:
            cell.text = ''
            cell.background_color = (0.5, 0.5, 0.5, 1) # 연한 회색으로 초기화
            cell.is_completed = False
        
        # 작업 번호 초기화
        for label in self.task_number_labels:
            label.text = ''
        
        self.complete_btn.disabled = True
        self.complete_btn.background_color = (0.5, 0.5, 0.5, 1) # 비활성화 시 어두운 회색 (변경)


class WarehouseApp(App):
    def build(self):
        self.title = '물류 창고 관리 시스템'
        return WarehouseGUI()


if __name__ == '__main__':
    WarehouseApp().run()