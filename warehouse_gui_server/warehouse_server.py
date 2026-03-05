#!/usr/bin/env python3
"""
창고 관리 서버 (노트북) - v2
- start_order 시점에 재고 차감 (동시 주문 방지)
- HTTP API: 피킹 리스트 제공
- MQTT: 주문 시작/완료 수신
"""
import sqlite3
import json
import threading
import time
from openpyxl import load_workbook
from flask import Flask, jsonify, request
# from flask_cors import CORS  # 필요시 설치: pip install flask-cors

# MQTT 라이브러리
try:
    import paho.mqtt.client as mqtt
except ImportError:
    print("⚠️  paho-mqtt 라이브러리가 필요합니다: pip install paho-mqtt")
    mqtt = None


class WarehouseServer:
    """창고 관리 서버"""
    
    def __init__(self, db_path='warehouse.db', mqtt_broker='localhost', mqtt_port=1883, http_port=5000):
        self.db_path = db_path
        self.mqtt_broker = mqtt_broker
        self.mqtt_port = mqtt_port
        self.http_port = http_port
        self.mqtt_client = None
        
        # 사용자별 현재 주문 추적
        self.current_orders = {}  # {user_id: order_number}
        
        # DB 연결 풀 (스레드 안전)
        self.db_lock = threading.Lock()
        
        # Flask 앱
        self.app = Flask(__name__)
        # CORS(self.app)  # 필요시 활성화
        self.setup_routes()
        
    def setup_routes(self):
        """Flask 라우트 설정"""
        
        @self.app.route('/api/picking/user/<int:user_id>/order/<int:order_number>', methods=['GET'])
        def get_picking_list(user_id, order_number):
            """피킹 리스트 요청"""
            picking_list = self.generate_picking_list(user_id, order_number)
            
            if picking_list:
                return jsonify({
                    'status': 'success',
                    '사용자ID': user_id,
                    '주문번호': order_number,
                    '피킹리스트': picking_list
                }), 200
            else:
                return jsonify({
                    'status': 'error',
                    'message': '주문을 찾을 수 없습니다'
                }), 404
        
        @self.app.route('/api/inventory/status', methods=['GET'])
        def get_inventory_status():
            """전체 재고 현황 조회"""
            status = self.get_inventory_status()
            return jsonify(status), 200
        
        @self.app.route('/health', methods=['GET'])
        def health_check():
            """서버 상태 확인"""
            return jsonify({'status': 'ok', 'service': 'warehouse_server'}), 200
    
    def init_mqtt(self):
        """MQTT 클라이언트 초기화"""
        if mqtt is None:
            print("❌ MQTT 라이브러리 없음")
            return False
        
        self.mqtt_client = mqtt.Client(client_id="warehouse_server")
        self.mqtt_client.on_connect = self.on_mqtt_connect
        self.mqtt_client.on_message = self.on_mqtt_message
        
        try:
            self.mqtt_client.connect(self.mqtt_broker, self.mqtt_port, 60)
            self.mqtt_client.loop_start()
            print(f"✅ MQTT 브로커 연결: {self.mqtt_broker}:{self.mqtt_port}")
            return True
        except Exception as e:
            print(f"❌ MQTT 연결 실패: {e}")
            return False
    
    def on_mqtt_connect(self, client, userdata, flags, rc):
        """MQTT 연결 콜백"""
        if rc == 0:
            print("✅ MQTT 브로커 연결 성공")
            # 토픽 구독
            client.subscribe("warehouse/order/start")
            client.subscribe("warehouse/shelf/complete")
            client.subscribe("warehouse/order/complete")
            print("📡 토픽 구독 완료")
        else:
            print(f"❌ MQTT 연결 실패: {rc}")
    
    def on_mqtt_message(self, client, userdata, msg):
        """MQTT 메시지 수신 콜백"""
        topic = msg.topic
        payload = json.loads(msg.payload.decode())
        
        print(f"\n📩 MQTT 수신: {topic}")
        print(f"   데이터: {payload}")
        
        if topic == "warehouse/order/start":
            self.handle_order_start(payload)
        elif topic == "warehouse/shelf/complete":
            self.handle_shelf_complete(payload)
        elif topic == "warehouse/order/complete":
            self.handle_order_complete(payload)
    
    def handle_order_start(self, data):
        """주문 시작 처리 - 즉시 재고 차감!"""
        user_id = data.get('사용자ID')
        order_number = data.get('주문번호')
        
        print(f"🚀 주문 시작: 사용자{user_id}, 주문번호{order_number}")
        
        # ⭐ 핵심: 주문 시작과 동시에 재고 차감!
        success = self.deduct_inventory(user_id, order_number)
        
        if success:
            print(f"✅ 재고 즉시 차감 완료 (예약)")
            print(f"   → 라즈베리파이는 HTTP API로 피킹 리스트 요청하세요.")
            
            # 현재 주문 추적
            self.current_orders[user_id] = order_number
        else:
            print(f"❌ 재고 부족! 주문을 처리할 수 없습니다.")
            # TODO: 라즈베리파이에 재고 부족 알림 (필요시)
    
    def handle_shelf_complete(self, data):
        """선반 피킹 완료 처리 - 로그만 기록"""
        user_id = data.get('사용자ID')
        shelf_number = data.get('선반번호')
        
        print(f"📦 선반 완료: 사용자{user_id}, 선반{shelf_number}")
    
    def handle_order_complete(self, data):
        """주문 완료 처리 - 로그만 기록 (재고는 이미 차감됨)"""
        user_id = data.get('사용자ID')
        order_number = data.get('주문번호')
        
        print(f"🎉 주문 완료: 사용자{user_id}, 주문번호{order_number}")
        print(f"   (재고는 주문 시작 시 이미 차감되었습니다)")
    
    def generate_picking_list(self, user_id, order_number):
        """
        주문번호에 해당하는 피킹 리스트 생성
        
        반환 형식:
        [
            {"선반번호": "1-1-1", "물건": "드롭스", "개수": 3},
            {"선반번호": "2-1-2", "물건": "퍼지", "개수": 2},
            ...
        ]
        """
        # order_file = f'사용자{user_id}주문.xlsx'
        order_file = f'../webots_simulation/Database/사용자{user_id}주문.xlsx'

        try:
            wb = load_workbook(order_file)
            ws = wb.active
            
            picking_list = []
            
            # 주문 데이터 읽기
            for row in ws.iter_rows(min_row=3, values_only=True):
                order_num = row[0]
                item = row[1]
                quantity = row[2]
                
                if order_num == order_number and item and quantity:
                    # DB에서 선반 번호 찾기
                    shelf_number = self.get_shelf_by_item(item)
                    
                    if shelf_number:
                        picking_list.append({
                            "선반번호": shelf_number,
                            "물건": item,
                            "개수": quantity
                        })
                    else:
                        print(f"⚠️  물건을 찾을 수 없음: {item}")
            
            # 피킹 경로 최적화 (구역 → 열 → 층 순으로 정렬)
            picking_list = self.optimize_picking_route(picking_list)
            
            return picking_list
            
        except FileNotFoundError:
            print(f"❌ 주문 파일 없음: {order_file}")
            return []
        except Exception as e:
            print(f"❌ 주문 로드 오류: {e}")
            return []
    
    def get_shelf_by_item(self, item):
        """물건명으로 선반번호 조회"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("SELECT shelf_number FROM inventory WHERE item = ?", (item,))
            result = cursor.fetchone()
            conn.close()
            
            return result[0] if result else None
    
    def optimize_picking_route(self, picking_list):
        """
        피킹 경로 최적화
        - 구역(zone) → 열(row) → 층(tier) 순으로 정렬
        """
        def parse_shelf(shelf_str):
            parts = shelf_str.split('-')
            return (int(parts[0]), int(parts[1]), int(parts[2]))
        
        # 선반번호 기준 정렬
        picking_list.sort(key=lambda x: parse_shelf(x['선반번호']))
        
        return picking_list
    
    def deduct_inventory(self, user_id, order_number):
        """
        주문 시작 시 재고 차감 (트랜잭션)
        
        ⭐ 동시 주문 방지:
        - 사용자1이 드롭스 3개 주문 시작 → 즉시 차감
        - 사용자2가 드롭스 5개 주문 시작 → 남은 재고만 사용 가능
        """
        # order_file = f'사용자{user_id}주문.xlsx'
        order_file = f'../webots_simulation/Database/사용자{user_id}주문.xlsx'

        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            try:
                # 트랜잭션 시작
                cursor.execute("BEGIN TRANSACTION")
                
                # 주문 데이터 읽기
                wb = load_workbook(order_file)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=3, values_only=True):
                    order_num = row[0]
                    item = row[1]
                    quantity = row[2]
                    
                    if order_num == order_number and item and quantity:
                        # 현재 재고 확인
                        cursor.execute("SELECT stock FROM inventory WHERE item = ?", (item,))
                        result = cursor.fetchone()
                        
                        if not result:
                            raise Exception(f"물건을 찾을 수 없음: {item}")
                        
                        current_stock = result[0]
                        
                        if current_stock < quantity:
                            raise Exception(f"재고 부족: {item} (필요: {quantity}, 현재: {current_stock})")
                        
                        # 재고 차감
                        cursor.execute("""
                            UPDATE inventory 
                            SET stock = stock - ? 
                            WHERE item = ?
                        """, (quantity, item))
                        
                        new_stock = current_stock - quantity
                        print(f"   ✓ {item}: {quantity}개 차감 → 남은 재고 {new_stock}")
                
                # 커밋
                conn.commit()
                conn.close()
                return True
                
            except Exception as e:
                cursor.execute("ROLLBACK")
                conn.close()
                print(f"❌ 재고 차감 실패: {e}")
                return False
    
    def get_inventory_status(self):
        """현재 재고 상태 조회"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 구역별 통계
        cursor.execute("""
            SELECT zone, COUNT(*), SUM(stock) 
            FROM inventory 
            GROUP BY zone 
            ORDER BY zone
        """)
        zone_stats = []
        for zone, items, total_stock in cursor.fetchall():
            zone_stats.append({
                'zone': zone,
                'items': items,
                'total_stock': total_stock
            })
        
        # 재고 부족 항목 (5개 이하)
        cursor.execute("SELECT item, stock FROM inventory WHERE stock <= 5 ORDER BY stock")
        low_stock = [{'item': item, 'stock': stock} for item, stock in cursor.fetchall()]
        
        conn.close()
        
        return {
            'zone_stats': zone_stats,
            'low_stock': low_stock
        }
    
    def show_inventory_status(self):
        """현재 재고 상태 출력 (콘솔용)"""
        status = self.get_inventory_status()
        
        print("\n" + "="*60)
        print("📦 현재 재고 상태")
        print("="*60)
        
        for stat in status['zone_stats']:
            print(f"구역 {stat['zone']}: {stat['items']}개 품목, 총 {stat['total_stock']}개")
        
        if status['low_stock']:
            print(f"\n⚠️  재고 부족 경고:")
            for item_info in status['low_stock']:
                print(f"   {item_info['item']}: {item_info['stock']}개")
    
    def run_flask(self):
        """Flask HTTP 서버 실행 (별도 스레드)"""
        print(f"✅ HTTP API 서버 시작: http://0.0.0.0:{self.http_port}")
        self.app.run(host='0.0.0.0', port=self.http_port, threaded=True, use_reloader=False)
    
    def run(self):
        """서버 실행"""
        print("="*60)
        print("🏭 창고 관리 서버 시작")
        print("="*60)
        
        # 재고 상태 표시
        self.show_inventory_status()
        
        # MQTT 초기화
        mqtt_ok = self.init_mqtt()
        if not mqtt_ok:
            print("⚠️  MQTT 없이 계속 실행 (HTTP API만 사용)")
        
        # Flask 서버 시작 (별도 스레드)
        flask_thread = threading.Thread(target=self.run_flask, daemon=True)
        flask_thread.start()
        
        print("\n✅ 서버 준비 완료")
        print(f"   - HTTP API: http://localhost:{self.http_port}")
        print(f"   - MQTT: {self.mqtt_broker}:{self.mqtt_port}")
        print("   (Ctrl+C로 종료)\n")
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n👋 서버 종료 중...")
            if self.mqtt_client:
                self.mqtt_client.loop_stop()
                self.mqtt_client.disconnect()


def main():
    """메인 함수"""
    server = WarehouseServer(
        db_path='warehouse.db',
        mqtt_broker='localhost',  # MQTT 브로커 주소
        mqtt_port=1883,
        http_port=5000  # HTTP API 포트
    )
    server.run()


if __name__ == '__main__':
    main()