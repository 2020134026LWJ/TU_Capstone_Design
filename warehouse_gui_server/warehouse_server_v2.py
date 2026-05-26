#!/usr/bin/env python3
"""
창고 관리 서버 (노트북) - v2
- shelf_complete 시점에 재고 차감 (사용자 GUI 터치 시 즉시)
- HTTP API: 피킹 리스트 제공
- MQTT: 주문 시작/완료 수신
- [테스트용] 선반 도착 신호 수동 전송 (콘솔 입력)
  → 추후 알고리즘 코드가 직접 shelf_arrived를 전송하면 서버 코드는 사용 안 함
"""
import sqlite3
import json
import threading
import time
from openpyxl import load_workbook
from flask import Flask, jsonify, request

# MQTT 라이브러리
try:
    import paho.mqtt.client as mqtt
except ImportError:
    print("⚠️  paho-mqtt 라이브러리가 필요합니다: pip install paho-mqtt")
    mqtt = None


class WarehouseServer:
    """창고 관리 서버"""
    
    def __init__(self, db_path='warehouse.db', mqtt_broker='localhost', mqtt_port=1883, http_port=5000):
        self.db_path = db_path
        self.mqtt_broker = mqtt_broker
        self.mqtt_port = mqtt_port
        self.http_port = http_port
        self.mqtt_client = None
        
        # 사용자별 현재 주문 추적
        self.current_orders = {}  # {user_id: order_number}

        # 사용자별 피킹리스트의 선반 목록 추적 (테스트용 선택지 제공)
        # {user_id: ["1-1", "2-3", ...]}
        self.user_shelf_groups = {}

        # 예약된 항목 추적 (미완료 시 reserved 해제용)
        # {user_id: [(item, quantity), ...]}
        self.pending_reservations = {}

        # 현재 도착하여 작업 중(미완료)인 선반 추적
        # {user_id: "1-1"} — 이 선반이 완료되기 전엔 다른 선반 호출 금지
        self.active_shelf = {}

        # DB 연결 풀 (스레드 안전)
        self.db_lock = threading.Lock()

        # Flask 앱
        self.app = Flask(__name__)
        self.setup_routes()
        self.init_db()
        
    def init_db(self):
        """DB 초기화: reserved 컬럼/진행상태 테이블 보강, 서버 시작 시 reserved 전체 초기화"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(inventory)")
            columns = [row[1] for row in cursor.fetchall()]
            if 'reserved' not in columns:
                cursor.execute("ALTER TABLE inventory ADD COLUMN reserved INTEGER NOT NULL DEFAULT 0")
                print("✅ DB: reserved 컬럼 추가 완료")

            # 피킹 진행 상태 테이블 (완료 품목 영구 저장) - 기존 DB에도 자동 추가
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS picking_progress (
                    user_id INTEGER NOT NULL,
                    order_number INTEGER NOT NULL,
                    shelf_number TEXT NOT NULL,
                    item TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    completed_at TEXT,
                    PRIMARY KEY (user_id, order_number, item)
                )
            """)
            # 사용자별 마지막 작업 주문번호 (자동 복원 재개 지점)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS user_state (
                    user_id INTEGER PRIMARY KEY,
                    current_order_number INTEGER NOT NULL DEFAULT 1
                )
            """)

            # 서버 재시작 시 이전 예약 상태 초기화 (진행 상태는 유지)
            cursor.execute("UPDATE inventory SET reserved = 0")
            conn.commit()
            conn.close()

    def setup_routes(self):
        """Flask 라우트 설정"""
        
        @self.app.route('/api/picking/user/<int:user_id>/order/<int:order_number>', methods=['GET'])
        def get_picking_list(user_id, order_number):
            picking_list = self.generate_picking_list(user_id, order_number)
            total_orders = self.get_total_orders(user_id)
            if picking_list:
                return jsonify({
                    'status': 'success',
                    '사용자ID': user_id,
                    '주문번호': order_number,
                    '총주문수': total_orders,
                    '피킹리스트': picking_list
                }), 200
            else:
                return jsonify({
                    'status': 'error',
                    'message': '주문을 찾을 수 없습니다'
                }), 404
        
        @self.app.route('/api/user/<int:user_id>/state', methods=['GET'])
        def get_user_state_route(user_id):
            current_order = self.get_user_state(user_id)
            return jsonify({
                'status': 'success',
                '사용자ID': user_id,
                '현재주문번호': current_order,
            }), 200

        @self.app.route('/api/inventory/status', methods=['GET'])
        def get_inventory_status():
            status = self.get_inventory_status()
            return jsonify(status), 200
        
        @self.app.route('/health', methods=['GET'])
        def health_check():
            return jsonify({'status': 'ok', 'service': 'warehouse_server'}), 200
    
    def init_mqtt(self):
        """MQTT 클라이언트 초기화"""
        if mqtt is None:
            print("❌ MQTT 라이브러리 없음")
            return False
        
        self.mqtt_client = mqtt.Client(client_id="warehouse_server")
        self.mqtt_client.on_connect = self.on_mqtt_connect
        self.mqtt_client.on_message = self.on_mqtt_message
        
        try:
            self.mqtt_client.connect(self.mqtt_broker, self.mqtt_port, 60)
            self.mqtt_client.loop_start()
            print(f"✅ MQTT 브로커 연결: {self.mqtt_broker}:{self.mqtt_port}")
            return True
        except Exception as e:
            print(f"❌ MQTT 연결 실패: {e}")
            return False
    
    def on_mqtt_connect(self, client, userdata, flags, rc):
        """MQTT 연결 콜백"""
        if rc == 0:
            print("✅ MQTT 브로커 연결 성공")
            client.subscribe("warehouse/order/start")
            client.subscribe("warehouse/shelf/complete")
            client.subscribe("warehouse/order/complete")
            client.subscribe("warehouse/order/all_complete")
            print("📡 토픽 구독 완료")
        else:
            print(f"❌ MQTT 연결 실패: {rc}")
    
    def on_mqtt_message(self, client, userdata, msg):
        """MQTT 메시지 수신 콜백"""
        topic = msg.topic
        payload = json.loads(msg.payload.decode())
        
        print(f"\n📩 MQTT 수신: {topic}")
        print(f"   데이터: {payload}")
        
        if topic == "warehouse/order/start":
            self.handle_order_start(payload)
        elif topic == "warehouse/shelf/complete":
            self.handle_shelf_complete(payload)
        elif topic == "warehouse/order/complete":
            self.handle_order_complete(payload)
        elif topic == "warehouse/order/all_complete":
            self.handle_all_orders_complete(payload)
    
    def handle_order_start(self, data):
        """주문 시작 처리 - 가상 차감(reserved)으로 이중 피킹 방지 + 선반 목록 준비"""
        user_id = data.get('사용자ID')
        order_number = data.get('주문번호')

        print(f"🚀 주문 시작: 사용자{user_id}, 주문번호{order_number}")

        # 이전 주문의 남은 예약 해제 (정상 흐름에선 이미 0이지만 안전 처리)
        self.release_reservation(user_id)

        # 가상 차감: 다른 사용자가 같은 물건을 동시에 못 가져가도록 예약
        self.reserve_inventory(user_id, order_number)

        self.current_orders[user_id] = order_number

        # 새 주문 시작 → 작업 중인 선반 없음
        self.active_shelf.pop(user_id, None)

        picking_list = self.generate_picking_list(user_id, order_number)
        shelf_groups = self.extract_shelf_groups(picking_list)
        self.user_shelf_groups[user_id] = shelf_groups

        print(f"   → 사용자{user_id} 선반 목록: {shelf_groups}")
        print(f"   → 콘솔에서 선반 도착 신호를 전송하세요 (명령어: s)")
    
    def handle_shelf_complete(self, data):
        """선반 피킹 완료 처리 - 터치 시 즉시 재고 차감 + 진행 상태 저장"""
        user_id = data.get('사용자ID')
        shelf_number = data.get('선반번호')
        items = data.get('품목목록', [])
        # 주문번호: 페이로드 우선, 없으면 현재 진행 중인 주문 사용
        order_number = data.get('주문번호', self.current_orders.get(user_id))
        print(f"📦 선반 완료: 사용자{user_id}, 선반{shelf_number}, 주문{order_number}")

        if items:
            self.deduct_stock(user_id, order_number, items)

        # 작업 중이던 선반의 모든 품목이 완료되면 해제 → 다음 선반 호출 허용
        active = self.active_shelf.get(user_id)
        if active is not None and self._shelf_group_completed(user_id, order_number, active):
            del self.active_shelf[user_id]
            print(f"✅ 선반 {active} 완료 → 다음 선반 호출 가능")

    def _shelf_group_completed(self, user_id, order_number, shelf_group):
        """해당 주문에서 선반 그룹(구역-열)의 모든 품목이 완료되었는지 확인"""
        picking_list = self.generate_picking_list(user_id, order_number)
        group_items = [
            p['물건'] for p in picking_list
            if '-'.join(p['선반번호'].split('-')[:2]) == shelf_group
        ]
        if not group_items:
            return True
        completed = self.get_completed_items(user_id, order_number)
        return all(it in completed for it in group_items)

    # ── 진행 상태(영구 저장) 헬퍼 ──────────────────────────────────────

    def get_user_state(self, user_id):
        """사용자별 마지막 작업 주문번호 조회 (없으면 1)"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT current_order_number FROM user_state WHERE user_id = ?", (user_id,)
            )
            result = cursor.fetchone()
            conn.close()
            return result[0] if result else 1

    def set_user_state(self, user_id, order_number):
        """사용자별 현재 주문번호 저장 (재개 지점)"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT OR REPLACE INTO user_state (user_id, current_order_number) VALUES (?, ?)",
                (user_id, order_number)
            )
            conn.commit()
            conn.close()

    def get_completed_items(self, user_id, order_number):
        """해당 주문에서 이미 완료(차감)한 품목명 집합 반환"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT item FROM picking_progress WHERE user_id = ? AND order_number = ?",
                (user_id, order_number)
            )
            items = {row[0] for row in cursor.fetchall()}
            conn.close()
            return items

    def clear_user_progress(self, user_id):
        """전체 주문 완료 시 그 사용자 진행 기록/상태 정리 (다음 사이클 대비)"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM picking_progress WHERE user_id = ?", (user_id,))
            cursor.execute("DELETE FROM user_state WHERE user_id = ?", (user_id,))
            conn.commit()
            conn.close()
            print(f"🧹 사용자{user_id} 진행 상태 정리 완료")

    def deduct_stock(self, user_id, order_number, items):
        """터치(shelf_complete) 시 stock 차감 + reserved 해제 + 진행 상태 저장

        이미 picking_progress에 기록된 품목은 재차감하지 않는다(이중 차감 방지).
        """
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            try:
                cursor.execute("BEGIN TRANSACTION")
                for item_info in items:
                    item = item_info.get('물건')
                    quantity = item_info.get('개수')
                    shelf_number = item_info.get('선반번호')

                    # 이미 완료 기록이 있으면 건너뜀 (재차감 방지)
                    if order_number is not None:
                        cursor.execute(
                            "SELECT 1 FROM picking_progress WHERE user_id = ? AND order_number = ? AND item = ?",
                            (user_id, order_number, item)
                        )
                        if cursor.fetchone():
                            print(f"   ↺ {item}: 이미 완료됨 → 차감 생략")
                            continue

                    # 선반번호 미제공 시 재고 테이블에서 보완
                    if not shelf_number:
                        cursor.execute("SELECT shelf_number FROM inventory WHERE item = ?", (item,))
                        r = cursor.fetchone()
                        shelf_number = r[0] if r else ''

                    cursor.execute(
                        "UPDATE inventory SET stock = stock - ?, reserved = MAX(0, reserved - ?) WHERE item = ?",
                        (quantity, quantity, item)
                    )
                    # 진행 상태 영구 저장
                    if order_number is not None:
                        cursor.execute(
                            """INSERT OR IGNORE INTO picking_progress
                               (user_id, order_number, shelf_number, item, quantity, completed_at)
                               VALUES (?, ?, ?, ?, ?, datetime('now', 'localtime'))""",
                            (user_id, order_number, shelf_number, item, quantity)
                        )
                    print(f"   ✓ {item}: {quantity}개 차감 (reserved 해제, 진행 저장)")
                conn.commit()
                print(f"✅ 사용자{user_id} 재고 차감/진행 저장 완료")
            except Exception as e:
                cursor.execute("ROLLBACK")
                print(f"❌ 재고 차감 실패: {e}")
            conn.close()

    def reserve_inventory(self, user_id, order_number):
        """주문 시작 시 가상 차감 - 가용 재고(stock-reserved) 기준으로 예약"""
        order_file = f'사용자{user_id}주문.xlsx'
        reserved_items = []

        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            try:
                cursor.execute("BEGIN TRANSACTION")

                # 재개 지점 저장
                cursor.execute(
                    "INSERT OR REPLACE INTO user_state (user_id, current_order_number) VALUES (?, ?)",
                    (user_id, order_number)
                )

                # 이미 완료(차감)된 품목은 다시 예약하지 않음
                cursor.execute(
                    "SELECT item FROM picking_progress WHERE user_id = ? AND order_number = ?",
                    (user_id, order_number)
                )
                completed = {row[0] for row in cursor.fetchall()}

                wb = load_workbook(order_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=3, values_only=True):
                    order_num = row[0]
                    item = row[1]
                    quantity = row[2]

                    if order_num == order_number and item and quantity:
                        if item in completed:
                            print(f"   ↺ {item}: 이미 완료됨 → 예약 생략")
                            continue

                        cursor.execute(
                            "SELECT stock, reserved FROM inventory WHERE item = ?", (item,)
                        )
                        result = cursor.fetchone()
                        if not result:
                            raise Exception(f"물건을 찾을 수 없음: {item}")

                        stock, reserved = result
                        available = stock - reserved
                        if available < quantity:
                            raise Exception(f"재고 부족: {item} (필요: {quantity}, 가용: {available})")

                        cursor.execute(
                            "UPDATE inventory SET reserved = reserved + ? WHERE item = ?",
                            (quantity, item)
                        )
                        reserved_items.append((item, quantity))
                        print(f"   ✓ {item}: {quantity}개 예약 (가용 재고 {available - quantity})")

                conn.commit()
                self.pending_reservations[user_id] = reserved_items
                print(f"✅ 사용자{user_id} 가상 차감 완료")
                self.mqtt_client.publish(
                    "warehouse/order/accepted",
                    json.dumps({"사용자ID": user_id, "주문번호": order_number})
                )
            except Exception as e:
                cursor.execute("ROLLBACK")
                print(f"❌ 가상 차감 실패: {e}")
            conn.close()

    def release_reservation(self, user_id):
        """주문 완료/취소 시 남은 reserved 해제 (안전 처리)"""
        items = self.pending_reservations.pop(user_id, [])
        if not items:
            return
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            try:
                cursor.execute("BEGIN TRANSACTION")
                for item, quantity in items:
                    cursor.execute(
                        "UPDATE inventory SET reserved = MAX(0, reserved - ?) WHERE item = ?",
                        (quantity, item)
                    )
                conn.commit()
                print(f"✅ 사용자{user_id} 예약 해제 완료")
            except Exception as e:
                cursor.execute("ROLLBACK")
                print(f"❌ 예약 해제 실패: {e}")
            conn.close()
    
    def handle_order_complete(self, data):
        """주문 완료 처리"""
        user_id = data.get('사용자ID')
        order_number = data.get('주문번호')
        print(f"🎉 주문 완료: 사용자{user_id}, 주문번호{order_number}")
        
        # 선반 목록/작업 중 선반 초기화
        if user_id in self.user_shelf_groups:
            del self.user_shelf_groups[user_id]
        self.active_shelf.pop(user_id, None)

    def handle_all_orders_complete(self, data):
        """전체 주문 완료 처리 - 혹시 남은 예약 해제"""
        user_id = data.get('사용자ID')
        total = data.get('총주문수')
        print(f"🏁 사용자{user_id} 전체 주문 완료! (총 {total}건)")

        self.release_reservation(user_id)
        self.clear_user_progress(user_id)

        if user_id in self.current_orders:
            del self.current_orders[user_id]
        self.active_shelf.pop(user_id, None)
    
    def extract_shelf_groups(self, picking_list):
        """피킹리스트에서 중복 없는 선반 그룹(구역-열) 목록 추출

        이미 완료된 품목의 선반은 도착시킬 필요가 없으므로 제외한다.
        (한 그룹에 미완료 품목이 하나라도 있으면 그 그룹은 포함)
        """
        seen = set()
        groups = []
        for item in picking_list:
            if item.get('완료'):
                continue  # 완료된 품목 → 선반 도착 불필요
            shelf = item['선반번호']
            group = '-'.join(shelf.split('-')[:2])  # "1-1-2" → "1-1"
            if group not in seen:
                seen.add(group)
                groups.append(group)
        return groups
    
    def send_shelf_arrived(self, user_id, shelf_group):
        """
        [테스트용] 선반 도착 신호 MQTT 전송
        → 추후 알고리즘 코드가 직접 이 메시지를 전송함
        """
        if not self.mqtt_client:
            print("❌ MQTT 미연결")
            return

        # 현재 작업 중(미완료)인 선반이 있으면 다른 선반 호출 거부
        active = self.active_shelf.get(user_id)
        if active is not None and active != shelf_group:
            print(f"⛔ 선반 {shelf_group} 호출 거부: 사용자{user_id}가 선반 {active} 작업 중입니다. "
                  f"(완료 후 호출 가능)")
            return

        msg = {
            "type": "shelf_arrived",
            "사용자ID": user_id,
            "선반번호": shelf_group
        }

        self.mqtt_client.publish(
            "warehouse/shelf/arrived",
            json.dumps(msg, ensure_ascii=False)
        )
        # 이 선반을 작업 중으로 표시 (완료 전까지 다른 선반 호출 차단)
        self.active_shelf[user_id] = shelf_group
        print(f"✅ shelf_arrived 전송: 사용자{user_id}, 선반{shelf_group}")
    
    def run_test_console(self):
        """
        [테스트용] 콘솔 입력으로 선반 도착 신호 수동 전송
        
        명령어:
          s  → 선반 도착 신호 전송 (사용자ID, 선반번호 입력)
          q  → 서버 종료
        """
        print("\n" + "="*60)
        print("🧪 테스트 콘솔 시작")
        print("  s : 선반 도착 신호 전송")
        print("  q : 종료")
        print("="*60)
        
        while True:
            try:
                cmd = input("\n명령어 입력 > ").strip().lower()
                
                if cmd == 'q':
                    print("👋 종료")
                    break
                
                elif cmd == 's':
                    # 현재 주문 중인 사용자 목록 표시
                    if not self.current_orders:
                        print("⚠️  현재 주문 중인 사용자가 없습니다.")
                        print("   → GUI에서 작업시작 버튼을 먼저 누르세요.")
                        continue
                    
                    print("\n현재 주문 중인 사용자:")
                    for uid, order_num in self.current_orders.items():
                        shelves = self.user_shelf_groups.get(uid, [])
                        print(f"  사용자{uid} - 주문{order_num}번 | 선반 목록: {shelves}")
                    
                    # 사용자 선택
                    uid_input = input("사용자ID 입력 (1 or 2): ").strip()
                    try:
                        user_id = int(uid_input)
                    except ValueError:
                        print("❌ 숫자를 입력하세요.")
                        continue
                    
                    if user_id not in self.current_orders:
                        print(f"❌ 사용자{user_id}는 현재 주문 중이 아닙니다.")
                        continue
                    
                    # 선반 선택
                    shelves = self.user_shelf_groups.get(user_id, [])
                    if not shelves:
                        print("⚠️  선반 목록이 없습니다.")
                        continue
                    
                    print(f"\n사용자{user_id} 선반 목록:")
                    for i, shelf in enumerate(shelves):
                        print(f"  {i+1}. {shelf}")
                    
                    shelf_input = input("선반 번호 선택 (번호 입력): ").strip()
                    try:
                        shelf_idx = int(shelf_input) - 1
                        if shelf_idx < 0 or shelf_idx >= len(shelves):
                            print("❌ 올바른 번호를 입력하세요.")
                            continue
                        shelf_group = shelves[shelf_idx]
                    except ValueError:
                        print("❌ 숫자를 입력하세요.")
                        continue
                    
                    # 전송
                    self.send_shelf_arrived(user_id, shelf_group)
                
                else:
                    print("❌ 알 수 없는 명령어입니다. (s: 선반 도착, q: 종료)")
            
            except (EOFError, KeyboardInterrupt):
                break
    
    def get_total_orders(self, user_id):
        """사용자 주문 파일의 총 주문 수 반환"""
        order_file = f'사용자{user_id}주문.xlsx'
        try:
            wb = load_workbook(order_file)
            ws = wb.active
            order_numbers = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                order_num = row[0]
                if order_num is not None:
                    order_numbers.add(order_num)
            return len(order_numbers)
        except Exception as e:
            print(f"❌ 총 주문 수 조회 실패: {e}")
            return 0

    def generate_picking_list(self, user_id, order_number):
        """주문번호에 해당하는 피킹 리스트 생성"""
        order_file = f'사용자{user_id}주문.xlsx'
        
        try:
            # 이미 완료한 품목 (자동 복원 시 완료 표시 + 재고 체크 생략)
            completed = self.get_completed_items(user_id, order_number)

            wb = load_workbook(order_file)
            ws = wb.active
            picking_list = []

            for row in ws.iter_rows(min_row=3, values_only=True):
                order_num = row[0]
                item = row[1]
                quantity = row[2]

                if order_num == order_number and item and quantity:
                    is_done = item in completed
                    shelf_number, available = self.get_shelf_by_item(item, user_id=user_id)
                    if shelf_number is None:
                        print(f"⚠️  물건을 찾을 수 없음: {item} → 주문 {order_number} 스킵")
                        return []
                    # 완료된 품목은 이미 차감됨 → 재고 부족 체크 건너뜀
                    elif not is_done and available < quantity:
                        print(f"⚠️  재고 부족: {item} (필요: {quantity}, 가용: {available}) → 주문 {order_number} 스킵")
                        return []
                    else:
                        picking_list.append({
                            "선반번호": shelf_number,
                            "물건": item,
                            "개수": quantity,
                            "완료": is_done
                        })

            picking_list = self.optimize_picking_route(picking_list)
            return picking_list
            
        except FileNotFoundError:
            print(f"❌ 주문 파일 없음: {order_file}")
            return []
        except Exception as e:
            print(f"❌ 주문 로드 오류: {e}")
            return []
    
    def get_shelf_by_item(self, item, user_id=None):
        """물건명으로 선반번호와 가용 재고 조회 (현재 사용자 예약분은 제외하고 계산)"""
        with self.db_lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT shelf_number, stock - COALESCE(reserved, 0) FROM inventory WHERE item = ?", (item,)
            )
            result = cursor.fetchone()
            conn.close()
            if not result:
                return (None, 0)
            shelf, available = result
            # 이 사용자가 이미 예약한 수량은 다시 더해서 원래 가용 재고로 복원
            if user_id is not None:
                user_rsv = {i: q for i, q in self.pending_reservations.get(user_id, [])}
                available += user_rsv.get(item, 0)
            return (shelf, available)
    
    def optimize_picking_route(self, picking_list):
        """피킹 경로 최적화 (구역 → 열 → 층 순)"""
        def parse_shelf(shelf_str):
            parts = shelf_str.split('-')
            return (int(parts[0]), int(parts[1]), int(parts[2]))
        picking_list.sort(key=lambda x: parse_shelf(x['선반번호']))
        return picking_list
    
    def get_inventory_status(self):
        """현재 재고 상태 조회"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT zone, COUNT(*), SUM(stock) 
            FROM inventory GROUP BY zone ORDER BY zone
        """)
        zone_stats = []
        for zone, items, total_stock in cursor.fetchall():
            zone_stats.append({'zone': zone, 'items': items, 'total_stock': total_stock})
        
        cursor.execute("SELECT item, stock FROM inventory WHERE stock <= 5 ORDER BY stock")
        low_stock = [{'item': item, 'stock': stock} for item, stock in cursor.fetchall()]
        
        conn.close()
        return {'zone_stats': zone_stats, 'low_stock': low_stock}
    
    def show_inventory_status(self):
        """현재 재고 상태 출력"""
        status = self.get_inventory_status()
        print("\n" + "="*60)
        print("📦 현재 재고 상태")
        print("="*60)
        for stat in status['zone_stats']:
            print(f"구역 {stat['zone']}: {stat['items']}개 품목, 총 {stat['total_stock']}개")
        if status['low_stock']:
            print(f"\n⚠️  재고 부족 경고:")
            for item_info in status['low_stock']:
                print(f"   {item_info['item']}: {item_info['stock']}개")
    
    def run_flask(self):
        """Flask HTTP 서버 실행 (별도 스레드)"""
        print(f"✅ HTTP API 서버 시작: http://0.0.0.0:{self.http_port}")
        self.app.run(host='0.0.0.0', port=self.http_port, threaded=True, use_reloader=False)
    
    def run(self):
        """서버 실행"""
        print("="*60)
        print("🏭 창고 관리 서버 시작")
        print("="*60)
        
        self.show_inventory_status()
        
        mqtt_ok = self.init_mqtt()
        if not mqtt_ok:
            print("⚠️  MQTT 없이 계속 실행 (HTTP API만 사용)")
        
        # Flask 서버 (별도 스레드)
        flask_thread = threading.Thread(target=self.run_flask, daemon=True)
        flask_thread.start()
        
        print("\n✅ 서버 준비 완료")
        print(f"   - HTTP API: http://localhost:{self.http_port}")
        print(f"   - MQTT: {self.mqtt_broker}:{self.mqtt_port}")
        
        # [테스트용] 콘솔 입력 루프 (메인 스레드에서 실행)
        self.run_test_console()
        
        # 종료
        if self.mqtt_client:
            self.mqtt_client.loop_stop()
            self.mqtt_client.disconnect()


def main():
    server = WarehouseServer(
        db_path='warehouse.db',
        mqtt_broker='localhost',
        mqtt_port=1883,
        http_port=5000
    )
    server.run()


if __name__ == '__main__':
    main()
